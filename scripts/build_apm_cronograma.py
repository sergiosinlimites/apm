#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Genera el cronograma reprogramado en APM.xlsx (hojas Cronograma y Ruta Crítica).

Cuenta todos los días calendario como hábiles salvo 4 festivos explícitos.
Sobrescribe el libro con fechas por fase, predecesoras y formato según paleta APM.
"""

from __future__ import annotations

import sys
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Any, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# --- Paleta proyecto APM ---
NAVY = "1E293B"
GRANATE = "A8201A"
FONDO_ALT = "F8FAFC"
WHITE = "FFFFFF"

HOLIDAYS = frozenset(
    {
        date(2026, 4, 1),
        date(2026, 4, 2),
        date(2026, 4, 3),
        date(2026, 5, 1),
    }
)

# Actividades marcadas como críticas en la especificación
CRITICAL_IDS = frozenset(
    {
        "3.2",
        "4.1",
        "5.5",
        "5.6",
        "6.4",
        "7.6",
        "8.2",
        "8.4",
        "9.1",
        "9.2",
        "10.4",
    }
)


def is_working_day(d: date) -> bool:
    """Retorna True si el día cuenta para el cronograma (no es festivo).

    Los fines de semana SÍ cuentan; solo se excluyen los días en ``HOLIDAYS``.
    """
    return d not in HOLIDAYS


def next_working_day(d: date) -> date:
    """Primer día en o después de ``d`` que no sea festivo."""
    cur = d
    while not is_working_day(cur):
        cur += timedelta(days=1)
    return cur


def add_working_days(start: date, extra: int) -> date:
    """Suma ``extra`` días hábiles (calendario sin festivos) a ``start`` (extra >= 0)."""
    if extra < 0:
        raise ValueError("extra debe ser >= 0")
    cur = start
    for _ in range(extra):
        cur += timedelta(days=1)
        cur = next_working_day(cur)
    return cur


def count_working_days(start: date, end: date) -> int:
    """Cuenta días hábiles (calendario sin festivos) inclusivos entre start y end."""
    if end < start:
        return 0
    n = 0
    cur = start
    while cur <= end:
        if is_working_day(cur):
            n += 1
        cur += timedelta(days=1)
    return n


def split_durations(total_wd: int, n: int) -> Tuple[int, ...]:
    """Reparte ``total_wd`` en ``n`` duraciones enteras >= 1 lo más equitativo posible."""
    if n < 1:
        raise ValueError("n debe ser >= 1")
    if total_wd < n:
        raise ValueError(f"total_wd ({total_wd}) debe ser >= n ({n}) para duraciones mínimas 1")
    base = total_wd // n
    rem = total_wd % n
    out = [base + 1] * rem + [base] * (n - rem)
    return tuple(out)


def schedule_sequential(
    phase_start: date,
    phase_end: date,
    durations: Sequence[int],
) -> List[Tuple[date, date]]:
    """Asigna inicio/fin secuencial por duraciones en días hábiles (sin festivos)."""
    cursor = next_working_day(phase_start)
    out: List[Tuple[date, date]] = []
    for dur in durations:
        if dur < 1:
            raise ValueError("duración mínima 1")
        s = cursor
        e = add_working_days(s, dur - 1)
        out.append((s, e))
        cursor = next_working_day(e + timedelta(days=1))
    return out


# --- Entregables: nombres exactos EDT (reprogramación APM) ---
ENTREGABLES: dict[str, str] = {
    "1.1": "Familia de bebidas y contexto operacional definidos",
    "1.2": "Portafolio de tres productos diferenciado",
    "1.3": "Especificación de envases y formatos establecida",
    "1.4": "Recetas y parámetros críticos documentados",
    "1.5": "Transformaciones térmicas/químicas identificadas",
    "1.6": "Estrategia de calidad y trazabilidad definida",
    "1.7": "Interacción operador–SCADA definida",
    "1.8": "Indicadores por producto establecidos",
    "2.1": "Levantamiento técnico del proceso consolidado",
    "2.2": "Diagrama general del proceso elaborado",
    "2.3": "Flujo de materiales e información definido",
    "2.4": "Estaciones, equipos y recursos caracterizados",
    "2.5": "Tiempos base de operación estimados",
    "2.6": "VSM estado actual elaborado",
    "2.7": "Desperdicios/Muda identificados",
    "2.8": "Indicadores AS-IS calculados",
    "2.9": "Fallas típicas y parámetros de mantenimiento caracterizados",
    "2.10": "Modelo conceptual de simulación documentado",
    "3.1": "Objetivos de mejora aprobados",
    "3.2": "Propuesta Lean/preautomatización formulada",
    "3.3": "VSM estado futuro elaborado",
    "3.4": "Estrategia multiproducto definida",
    "3.5": "Estrategia de cambio de formato definida",
    "3.6": "Estrategia de fallas y recuperación definida",
    "3.7": "Estrategia de rechazo y reproceso definida",
    "4.1": "Arquitectura ISA-95 documentada",
    "4.2": "Arquitectura de comunicaciones definida",
    "4.3": "Topología de red industrial definida",
    "4.4": "Lista de E/S y tags consolidada",
    "4.5": "PFD/P&ID del sistema elaborados",
    "4.6": "Matriz de señales y alarmas elaborada",
    "4.7": "Requisitos de historización, eventos y lotes definidos",
    "5.1": "Modelo AS-IS implementado",
    "5.2": "Modelo AS-IS validado",
    "5.3": "Escenarios AS-IS simulados",
    "5.4": "Desempeño AS-IS cuantificado",
    "5.5": "Escenarios TO-BE diseñados",
    "5.6": "Modelo TO-BE simulado",
    "5.7": "OEE y KPIs por escenario calculados",
    "5.8": "Comparativo AS-IS vs TO-BE consolidado",
    "6.1": "EDT y diccionario EDT aprobados",
    "6.2": "Cronograma base del proyecto estructurado",
    "6.3": "Plan de recursos consolidado",
    "6.4": "Presupuesto CAPEX/OPEX consolidado",
    "6.5": "Evidencias de precios recopiladas",
    "6.6": "Flujo de caja documentado",
    "6.7": "Indicadores financieros calculados",
    "6.8": "Sensibilidad financiera elaborada",
    "6.9": "Propuesta de valor formulada",
    "6.10": "Oferta comercial estructurada",
    "6.11": "Modelo de negocio resumido definido",
    "7.1": "Justificación de robotización sustentada",
    "7.2": "Robot y herramienta de toma seleccionados",
    "7.3": "Layout de celda elaborado",
    "7.4": "Sensores y periféricos de celda definidos",
    "7.5": "Tiempos de celda cuantificados",
    "7.6": "Simulación RobotStudio validada",
    "7.7": "Medidas de seguridad de celda definidas",
    "7.8": "Riesgos de celda evaluados",
    "8.1": "Filosofía de control definida",
    "8.2": "Grafcet del sistema elaborado",
    "8.3": "Interlocks y permisos definidos",
    "8.4": "Lógica Ladder implementada",
    "8.5": "Gestión de recetas implementada",
    "8.6": "Manejo de fallas implementado",
    "8.7": "Validación funcional del control completada",
    "9.1": "Modelo 3D del sistema desarrollado",
    "9.2": "Sensores y actuadores virtuales integrados",
    "9.3": "Integración con Logix Emulate implementada",
    "9.4": "Pruebas integradas ejecutadas",
    "9.5": "Validación funcional del sistema completada",
    "10.1": "Estándar HMI ISA-101 definido",
    "10.2": "Pantallas HMI diseñadas",
    "10.3": "Modelo de tags SCADA estructurado",
    "10.4": "SCADA en Ignition implementado",
    "10.5": "Comunicación OPC operativa",
    "10.6": "Alarmas, eventos e historización configurados",
    "10.7": "Interacción del operador habilitada",
    "10.8": "Validación funcional con gemelo digital completada",
    "11.1": "Criterios de aceptación definidos",
    "11.2": "Pruebas integrales end-to-end ejecutadas",
    "11.3": "Cuadro resumen final consolidado",
    "11.4": "Repositorio GitHub curado",
    "11.5": "Sitio web del proyecto completado",
    "11.6": "Video de proyecto producido",
    "11.7": "Sustentación intermedia preparada",
    "11.8": "Sustentación final preparada",
    "11.9": "Reflexiones de aprendizaje consolidadas",
}

PRED: dict[str, str] = {
    "1.1": "",
    "1.2": "1.1",
    "1.3": "1.2",
    "1.4": "1.3",
    "1.5": "1.4",
    "1.6": "1.4, 1.5",
    "1.7": "1.6",
    "1.8": "1.1",
    "2.1": "1.3",
    "2.2": "2.1",
    "2.3": "2.1",
    "2.4": "2.1",
    "2.5": "2.1",
    "2.6": "2.3, 2.5",
    "2.7": "2.6",
    "2.8": "2.5",
    "2.9": "2.4",
    "2.10": "2.4, 2.5",
    "3.1": "2.7",
    "3.2": "3.1",
    "3.3": "3.2",
    "3.4": "3.1, 3.2, 3.3",
    "3.5": "3.4",
    "3.6": "2.9, 3.1",
    "3.7": "3.6",
    "4.1": "3.7",
    "4.2": "4.1",
    "4.3": "4.2",
    "4.4": "4.1, 4.2, 4.3",
    "4.5": "4.1",
    "4.6": "4.4",
    "4.7": "4.4, 4.6",
    "5.1": "2.10",
    "5.2": "5.1",
    "5.3": "5.2",
    "5.4": "5.3",
    "5.5": "3.7, 5.4",
    "5.6": "5.5",
    "5.7": "5.3, 5.6",
    "5.8": "5.4, 5.7",
    "6.1": "1.8",
    "6.2": "6.1",
    "6.3": "6.2",
    "6.4": "6.3",
    "6.5": "6.4",
    "6.6": "6.4, 6.5",
    "6.7": "6.6",
    "6.8": "6.7",
    "6.9": "3.7, 6.8",
    "6.10": "6.9",
    "6.11": "6.10",
    "7.1": "2.4, 3.7",
    "7.2": "7.1",
    "7.3": "7.2",
    "7.4": "7.3",
    "7.5": "7.3, 7.4",
    "7.6": "7.2, 7.3, 7.4, 7.5",
    "7.7": "7.3",
    "7.8": "7.6",
    "8.1": "4.7",
    "8.2": "8.1",
    "8.3": "8.2",
    "8.4": "4.4, 8.2, 8.3",
    "8.5": "1.4, 8.4",
    "8.6": "2.9, 8.4",
    "8.7": "8.4, 8.5, 8.6",
    "9.1": "4.7",
    "9.2": "9.1",
    "9.3": "8.7, 9.2",
    "9.4": "9.3",
    "9.5": "9.4",
    "10.1": "1.7, 4.6",
    "10.2": "10.1",
    "10.3": "4.4, 10.2",
    "10.4": "10.2, 10.3",
    "10.5": "8.7, 10.4",
    "10.6": "4.6, 10.4",
    "10.7": "10.1",
    "10.8": "9.5, 10.4, 10.5, 10.6, 10.7",
    "11.1": "3.7, 5.8",
    "11.2": "8.7, 9.5, 10.8",
    "11.3": "5.8, 6.8",
    "11.4": "11.3",
    "11.5": "11.3",
    "11.6": "11.3",
    "11.7": "6.11",
    "11.8": "11.2, 11.3",
    "11.9": "11.8",
}


@dataclass
class PhaseBlock:
    """Bloque de fase para el cronograma."""

    key: str
    title: str
    start: date
    end: date
    ids: Tuple[str, ...]


def obs_for_item(
    aid: str,
    start: Optional[date],
    end: Optional[date],
    phase_end: date,
) -> str:
    """Genera observaciones: crítica y desbordes de ventana."""
    parts: List[str] = []
    if aid in CRITICAL_IDS:
        parts.append("Actividad crítica")
    if end is not None and end > phase_end:
        parts.append("Fecha fin posterior al cierre oficial de fase (suma de duraciones)")
    return "; ".join(parts)


def build_phase_blocks() -> List[PhaseBlock]:
    """Define ventanas por fase; las duraciones se reparten con split_durations."""
    specs: List[Tuple[str, str, date, date, Tuple[str, ...]]] = [
        (
            "F1",
            "Fase 1 — Alcance técnico del proceso definido",
            date(2026, 2, 11),
            date(2026, 2, 27),
            tuple(f"1.{i}" for i in range(1, 9)),
        ),
        (
            "F2",
            "Fase 2 — Levantamiento de información y análisis AS-IS",
            date(2026, 2, 13),
            date(2026, 3, 10),
            ("2.1", "2.2", "2.3", "2.4", "2.5", "2.6", "2.7", "2.8", "2.9", "2.10"),
        ),
        (
            "F6",
            "Fase 6 — Viabilidad técnico-económica sustentada",
            date(2026, 3, 4),
            date(2026, 3, 22),
            tuple(f"6.{i}" for i in range(1, 12)),
        ),
        (
            "F3",
            "Fase 3 — Diseño futuro del proceso definido",
            date(2026, 3, 10),
            date(2026, 3, 18),
            ("3.1", "3.2", "3.3", "3.4", "3.5", "3.6", "3.7"),
        ),
        (
            "F4",
            "Fase 4 — Arquitectura de automatización e instrumentación definida",
            date(2026, 3, 28),
            date(2026, 4, 7),
            tuple(f"4.{i}" for i in range(1, 8)),
        ),
        (
            "F5",
            "Fase 5 — Evaluación productiva por simulación completada",
            date(2026, 3, 28),
            date(2026, 4, 9),
            tuple(f"5.{i}" for i in range(1, 9)),
        ),
        (
            "F7",
            "Fase 7 — Celda robotizada diseñada y validada",
            date(2026, 3, 28),
            date(2026, 4, 7),
            tuple(f"7.{i}" for i in range(1, 9)),
        ),
        (
            "F8",
            "Fase 8 — Sistema de control industrial validado",
            date(2026, 4, 8),
            date(2026, 4, 16),
            tuple(f"8.{i}" for i in range(1, 8)),
        ),
        (
            "F9",
            "Fase 9 — Gemelo digital integrado validado",
            date(2026, 4, 17),
            date(2026, 4, 25),
            tuple(f"9.{i}" for i in range(1, 6)),
        ),
        (
            "F10",
            "Fase 10 — Supervisión SCADA implementada",
            date(2026, 4, 22),
            date(2026, 5, 9),
            tuple(f"10.{i}" for i in range(1, 9)),
        ),
    ]
    return [PhaseBlock(k, t, s, e, ids) for k, t, s, e, ids in specs]


def phase_durations(pb: PhaseBlock) -> Tuple[int, ...]:
    """Duraciones por ítem que suman los días hábiles de la ventana de fase."""
    n = len(pb.ids)
    total = count_working_days(pb.start, pb.end)
    return split_durations(total, n)


def schedule_f11_blocks() -> List[dict[str, Any]]:
    """Fase 11: integración, cierre e ítems con ventanas explícitas (sin 11.7)."""
    blocks: List[dict[str, Any]] = []
    int_start, int_end = date(2026, 5, 10), date(2026, 5, 15)
    clo_start, clo_end = date(2026, 5, 10), date(2026, 5, 19)
    wd_int = count_working_days(int_start, int_end)
    durs_int = split_durations(wd_int, 2)
    spans_int = schedule_sequential(int_start, int_end, durs_int)

    blocks.append(
        {
            "type": "phase",
            "id": "F11a",
            "title": "Fase 11 — Integración (11.1–11.2)",
            "start": int_start,
            "end": int_end,
        }
    )
    for aid, (s, e), dur in zip(("11.1", "11.2"), spans_int, durs_int):
        blocks.append(
            {
                "type": "item",
                "id": aid,
                "fase": "Fase 11 — Integración (11.1–11.2)",
                "start": s,
                "end": e,
                "dur": dur,
            }
        )

    wd_clo = count_working_days(clo_start, clo_end)
    durs_clo = split_durations(wd_clo, 4)
    spans_clo = schedule_sequential(clo_start, clo_end, durs_clo)

    blocks.append(
        {
            "type": "phase",
            "id": "F11b",
            "title": "Fase 11 — Cierre (11.3–11.6)",
            "start": clo_start,
            "end": clo_end,
        }
    )
    for aid, (s, e), dur in zip(("11.3", "11.4", "11.5", "11.6"), spans_clo, durs_clo):
        blocks.append(
            {
                "type": "item",
                "id": aid,
                "fase": "Fase 11 — Cierre (11.3–11.6)",
                "start": s,
                "end": e,
                "dur": dur,
            }
        )
    return blocks


def inspect_existing_xlsx(path: Path) -> None:
    """Imprime estructura del libro existente (inspección)."""
    print(f"\n=== Inspección: {path} ===")
    if not path.exists():
        print("Archivo no encontrado.")
        return
    wb = load_workbook(path, read_only=True, data_only=True)
    print("Hojas:", ", ".join(wb.sheetnames))
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n--- Hoja: {name} (filas={ws.max_row}, cols={ws.max_column}) ---")
        for i, row in enumerate(
            ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 60), values_only=True),
            start=1,
        ):
            print(i, row)
    wb.close()


def print_first_rows(path: Path, sheet: str, n: int = 30) -> None:
    """Imprime las primeras ``n`` filas de una hoja para verificación rápida."""
    if not path.exists():
        print(f"No existe {path}")
        return
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        print(f"Hoja '{sheet}' no encontrada.")
        wb.close()
        return
    ws = wb[sheet]
    print(f"\n=== Primeras {n} filas: {sheet} ===")
    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, n), values_only=True),
        start=1,
    ):
        print(i, row)
    wb.close()


def format_pred(value: str) -> str:
    """Formatea predecesoras vacías como em dash tipográfico."""
    if value is None or str(value).strip() == "":
        return "—"
    return str(value).strip()


def thin_border() -> Border:
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)


def autofit_columns(ws: Any, max_col: int) -> None:
    """Ajusta ancho de columnas de forma heurística."""
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 55)


def write_cronograma(ws: Worksheet) -> int:
    """Escribe hoja Cronograma. Retorna número de filas de datos (sin encabezado)."""
    headers = [
        "ID",
        "Fase",
        "Entregable",
        "Duración (días)",
        "Fecha Inicio",
        "Fecha Fin",
        "Predecesoras",
        "Observaciones",
    ]
    header_font = Font(bold=True, color=WHITE)
    header_fill = PatternFill("solid", fgColor=NAVY)
    granate_fill = PatternFill("solid", fgColor=GRANATE)
    granate_font = Font(bold=True, color=WHITE)
    alt_fill = PatternFill("solid", fgColor=FONDO_ALT)
    white_fill = PatternFill("solid", fgColor=WHITE)

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin_border()

    row_idx = 2
    data_rows = 0

    def write_phase_row(
        pid: str,
        fase_title: str,
        d_start: date,
        d_end: date,
    ) -> None:
        nonlocal row_idx, data_rows
        ws.cell(row=row_idx, column=1, value=pid)
        ws.cell(row=row_idx, column=2, value=fase_title)
        ws.cell(row=row_idx, column=3, value="(Resumen de fase)")
        ddays = count_working_days(d_start, d_end)
        ws.cell(row=row_idx, column=4, value=ddays)
        ws.cell(row=row_idx, column=5, value=d_start)
        ws.cell(row=row_idx, column=6, value=d_end)
        ws.cell(row=row_idx, column=7, value="—")
        ws.cell(row=row_idx, column=8, value="Ventana oficial de la fase")
        for c in range(1, 9):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = granate_fill
            cell.font = granate_font
            cell.border = thin_border()
            if c in (5, 6):
                cell.number_format = "DD-MM-YYYY"
        row_idx += 1
        data_rows += 1

    def write_item_row(
        aid: str,
        fase_title: str,
        entregable: str,
        dur: int,
        d_start: date,
        d_end: date,
        pred: str,
        obs: str,
        stripe_white: bool,
    ) -> None:
        nonlocal row_idx, data_rows
        fill = white_fill if stripe_white else alt_fill
        ws.cell(row=row_idx, column=1, value=aid)
        ws.cell(row=row_idx, column=2, value=fase_title)
        ws.cell(row=row_idx, column=3, value=entregable)
        ws.cell(row=row_idx, column=4, value=dur)
        ws.cell(row=row_idx, column=5, value=d_start)
        ws.cell(row=row_idx, column=6, value=d_end)
        ws.cell(row=row_idx, column=7, value=format_pred(pred))
        ws.cell(row=row_idx, column=8, value=obs)
        for c in range(1, 9):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c in (5, 6):
                cell.number_format = "DD-MM-YYYY"
        row_idx += 1
        data_rows += 1

    blocks = build_phase_blocks()
    phase_by_key = {b.key: b for b in blocks}

    def emit_phase_block(key: str) -> None:
        nonlocal stripe
        pb = phase_by_key[key]
        durs = phase_durations(pb)
        write_phase_row(pb.key, pb.title, pb.start, pb.end)
        spans = schedule_sequential(pb.start, pb.end, durs)
        for aid, (s, e), dur in zip(pb.ids, spans, durs):
            o = obs_for_item(aid, s, e, pb.end)
            write_item_row(
                aid,
                pb.title,
                ENTREGABLES[aid],
                dur,
                s,
                e,
                PRED[aid],
                o,
                stripe,
            )
            stripe = not stripe

    # Orden cronológico por inicio: F1, F2, F6, F3 → hitos marzo + 11.7 → F4, F5, F7 → …
    stripe = False
    for key in ("F1", "F2", "F6", "F3"):
        emit_phase_block(key)

    # Hitos intermedios (22–27 mar) + sustentación intermedia EDT 11.7
    march_milestones: List[Tuple[str, str, str, date, date, str]] = [
        (
            "H-CMP",
            "Hitos del proyecto",
            "Consolidación y empaquetado",
            date(2026, 3, 22),
            date(2026, 3, 23),
            "Cierra entregables de fases 1–3 y 6 previos a hitos de sustentación",
        ),
        (
            "H-AFN",
            "Hitos del proyecto",
            "Afinación para sustentación",
            date(2026, 3, 24),
            date(2026, 3, 27),
            "Preparación de material y ensayos",
        ),
        (
            "H-INT",
            "Hitos del proyecto",
            "Sustentación intermedia",
            date(2026, 3, 25),
            date(2026, 3, 27),
            "Ventana de sustentación intermedia",
        ),
    ]
    for mid, fase_t, title, ds, de, obs in march_milestones:
        write_item_row(
            mid,
            fase_t,
            title,
            count_working_days(ds, de),
            ds,
            de,
            "",
            obs,
            stripe,
        )
        stripe = not stripe

    s117, e117 = date(2026, 3, 25), date(2026, 3, 27)
    write_item_row(
        "11.7",
        "Fase 11 — Sustentación intermedia (EDT 11.7)",
        ENTREGABLES["11.7"],
        count_working_days(s117, e117),
        s117,
        e117,
        PRED["11.7"],
        obs_for_item("11.7", s117, e117, e117),
        stripe,
    )
    stripe = not stripe

    for key in ("F4", "F5", "F7", "F8", "F9", "F10"):
        emit_phase_block(key)

    for piece in schedule_f11_blocks():
        if piece["type"] == "phase":
            write_phase_row(
                piece["id"],
                piece["title"],
                piece["start"],
                piece["end"],
            )
        else:
            aid = piece["id"]
            pe = piece["end"]
            o = obs_for_item(aid, piece["start"], pe, pe)
            write_item_row(
                aid,
                piece["fase"],
                ENTREGABLES[aid],
                piece["dur"],
                piece["start"],
                pe,
                PRED[aid],
                o,
                stripe,
            )
            stripe = not stripe

    s118, e118 = date(2026, 5, 20), date(2026, 5, 22)
    write_item_row(
        "11.8",
        "Fase 11 — Sustentación final (EDT 11.8)",
        ENTREGABLES["11.8"],
        count_working_days(s118, e118),
        s118,
        e118,
        PRED["11.8"],
        obs_for_item("11.8", s118, e118, e118),
        stripe,
    )
    stripe = not stripe

    milestones_fin: List[Tuple[str, str, str, date, date, str]] = [
        (
            "H-FIN",
            "Hitos del proyecto",
            "Sustentación final",
            date(2026, 5, 20),
            date(2026, 5, 22),
            "Alineado con actividad 11.8",
        ),
    ]
    for mid, fase_t, title, ds, de, obs in milestones_fin:
        write_item_row(
            mid,
            fase_t,
            title,
            count_working_days(ds, de),
            ds,
            de,
            "",
            obs,
            stripe,
        )
        stripe = not stripe

    s119, e119 = date(2026, 5, 23), date(2026, 5, 27)
    write_item_row(
        "11.9",
        "Fase 11 — Reflexiones (EDT 11.9)",
        ENTREGABLES["11.9"],
        count_working_days(s119, e119),
        s119,
        e119,
        PRED["11.9"],
        obs_for_item("11.9", s119, e119, e119),
        stripe,
    )
    stripe = not stripe

    write_item_row(
        "T-REFL",
        "Hitos del proyecto",
        "Reflexiones y cierre",
        count_working_days(s119, e119),
        s119,
        e119,
        "",
        "Actividad 11.9 — reflexiones de aprendizaje",
        stripe,
    )

    ws.freeze_panes = "A2"
    for r in range(1, row_idx):
        ws.row_dimensions[r].height = None
    autofit_columns(ws, 8)
    return data_rows


def write_ruta_critica(wb: Workbook) -> int:
    """Hoja Ruta Crítica con textos y tabla de festivos."""
    ws = wb.create_sheet("Ruta Crítica")
    ws.cell(row=1, column=1, value="Rutas críticas y calendario")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color=NAVY)

    ws.cell(row=3, column=1, value="Ruta crítica intermedia")
    ws.cell(row=3, column=1).font = Font(bold=True)
    ws.cell(row=4, column=1, value="H-F1 → H-F2 → H-F3 → H-CMP → H-AFN → H-INT")
    ws.cell(row=5, column=1, value="(H-F1, H-F2, H-F3 representan cierres de fases 1–3 en secuencia lógica)")

    ws.cell(row=7, column=1, value="Ruta crítica final")
    ws.cell(row=7, column=1).font = Font(bold=True)
    ws.cell(row=8, column=1, value="T-Arq → T-PLC → T-SCADA → T-Int → H-Fin (ventana según cronograma)")
    ws.cell(row=9, column=1, value="Donde T-Arq = arquitectura (Fase 4), T-PLC = control (Fase 8), T-SCADA = SCADA (Fase 10), T-Int = integración (11.1–11.2), H-Fin = sustentación final.")

    ws.cell(row=11, column=1, value="Días no hábiles considerados en el cronograma")
    ws.cell(row=11, column=1).font = Font(bold=True)
    ws.cell(row=12, column=1, value="Semana Santa: 01-abr-2026 al 03-abr-2026")
    ws.cell(row=13, column=1, value="Día del Trabajo: 01-may-2026")
    ws.cell(row=14, column=1, value="Sábados y domingos SÍ se cuentan; solo se excluyen los festivos anteriores.")

    autofit_columns(ws, 3)
    return ws.max_row or 0


def build_workbook() -> Tuple[Workbook, int]:
    """Construye el libro completo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma"
    rows = write_cronograma(ws)
    write_ruta_critica(wb)
    return wb, rows


def main() -> int:
    root = Path(__file__).resolve().parents[1]
    target = root / "APM.xlsx"

    wb, nrows = build_workbook()
    wb.save(target)
    print("\n=== Resumen ===")
    print(f"Archivo guardado: {target}")
    print(f"Hojas: {', '.join(wb.sheetnames)}")
    cron = wb["Cronograma"]
    print(f"Hoja 'Cronograma': {cron.max_row} filas (incl. encabezado), {cron.max_column} columnas")
    print(f"Filas de contenido (aprox. datos): {nrows}")
    rc = wb["Ruta Crítica"]
    print(f"Hoja 'Ruta Crítica': {rc.max_row} filas usadas, {rc.max_column} columnas")

    print_first_rows(target, "Cronograma", 30)
    return 0


if __name__ == "__main__":
    sys.exit(main())
